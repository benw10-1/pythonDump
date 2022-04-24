import datetime
import sys
from openpyxl import load_workbook, Workbook
from tkinter import Tk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename

date = datetime.datetime.today()


class student:
    def __init__(self, index, name, messagecount=None):
        self.index = index
        self.name = name
        self.messagecount = messagecount


def writetosheet(array, sheet):
    print(array)
    sheet.delete_rows(idx=1, amount=len(array[0])-1)
    sheet.delete_cols(idx=1, amount=len(array)-1)
    sheet.insert_rows(idx=1, amount=len(array[0])-1)
    sheet.insert_cols(idx=1, amount=len(array)-1)

    for i, column in enumerate(sheet.iter_cols()):
        for p, row in enumerate(column):
            row.value = array[i][p-1]


def readtxt(filename):
    with open(filename, "r", encoding="utf-8") as file:
        count = 0
        acount = 0
        studentlist = {}
        messagecount = {}
        # reads content and puts into list separated by lines
        contents = file.readlines()
        for x in contents:
            if x:
                x = x.split()
                if "From" in x:
                    if "to" not in x:
                        name = " ".join(x).split(":")[2].split("From")[1].strip().lower()
                        message = " ".join(x).split(":")[3].strip()
                        if name not in studentlist or name not in messagecount:
                            messagecount[name] = 1
                            studentlist[name] = [message]

                        else:
                            messagecount[name] = messagecount[name] + 1
                            studentlist[name].append(message)

                    else:
                        name = " ".join(x).split(":")[2].split("From")[1].split("to")[0].strip().lower()
                        message = " ".join(x).split(":")[3].strip()
                        if name not in studentlist or name not in messagecount:
                            messagecount[name] = 1
                            studentlist[name] = [message]

                        else:
                            messagecount[name] = messagecount[name] + 1
                            studentlist[name].append(message)
    return messagecount, studentlist


def readsheet(sheet):
    sheetdata = []
    overwrite = False
    # iterate through each column
    for p, x in enumerate(sheet.iter_cols(values_only=True)):
        sheetdata.append([])
        # iterate through each row of column
        for i, y in enumerate(x):
            if i == 0:
                if y == "datetime.date(2020)":
                    yesorno = messagebox.askyesno("Override Warning",
                                                  "Would you like to override the data for {}?".format(datetime.date))
                    if yesorno == "No":
                        sys.exit()
                    else:
                        overwrite = True
                else:
                    sheetdata[p].append(y)
            else:
                sheetdata[p].append(y)

    # returns sheet data in data[coulums][rows] form
    return sheetdata, overwrite


def main():
    Tk().withdraw()
    newfile = False
    # ask for text file to process
    while True:
        sheetsfile = askopenfilename(filetypes=[("Excell files", "*.xlsx")], title="Select Sheet To Write To")
        if sheetsfile:
            break
        else:
            msgbox = messagebox.showwarning(title="Create new file", message="Are you sure you want to create a new file?")
            if msgbox:
                newfile = True
                break
            else:
                newfile = False

    textfile = askopenfilename(filetypes=[("Text files", "*.txt")], title="Select Zoom Message Log")
    # ask for Excell file to process

    if not textfile:
        messagebox.showerror(title="No file entry", message="No text file selected")
        sys.exit()

    if not newfile:
        students = {}
        studentsindex = {}
        fullstudents = []

        read = load_workbook(filename=sheetsfile)
        # sheet object to pass to readsheet function
        sheet = read.active

        # data variables
        sheetdata, overwrite = readsheet(sheet)
        # dictionary keys are student names
        messagecount, messagelist = readtxt(textfile)

        # names on sheet
        sheetnames = sheetdata[0]

        for i, x in enumerate(sheetdata):
            for o, y in enumerate(x):
                if o != 0:
                    if i == 0:
                        students[y] = []
                        studentsindex[y] = o
                    else:
                        students[y].append(sheetdata[i][studentsindex])

        # position of each student on the sheet

        for i, x in enumerate(sheetnames):
            if i != 0:
                if x not in messagecount:
                    messagecount[x] = None

        i = 0
        for i, x in enumerate(messagecount):
            if i != 0:
                complist.append((messagecount[x], x))

        addedrow = [None] * len(sheetindex)
        firstrow = [None] * (len(sheetindex)+1)
        for x in objlist:
            try:
                addedrow[sheetindex[x]-1] = messagecount[x.lower()]
            except Exception as e:
                addedrow[sheetindex[x]-1] = None

            firstrow[sheetindex[x]] = x
        addedrow[0] = date

        if overwrite:
            sheetdata[len(sheetdata)-1] = addedrow

        else:
            sheetdata.append(addedrow)

        sheetdata[0] = firstrow

        writetosheet(sheetdata, sheet)

        read.save(filename=sheetsfile)

        with open((str(datetime.datetime.now()).replace(".", ":").replace(":", "-") + "_messagelog.txt"),
                  "w+") as file2:
            sortlist = [x for x in messagelist]
            sortlist.sort()
            for y in sortlist:
                count = 1
                splitter = y.split(" ")
                uppered = []
                for p in splitter:
                    uppered.append(p[0].upper() + p[1:])
                file2.writelines(" ".join(uppered) + ":" + "\n")
                for x in messagelist[y]:
                    file2.writelines((" " * (len(y) + 1)) + str(count) + ". " + x + "\n")
                    count += 1
            file2.close()

    else:
        workbook = Workbook()
        sheet = workbook.active

        messagecount, messagelist = readtxt(textfile)

        sortlist = [x for x in messagecount]
        sortlist.sort()
        upperedar = []
        for y in sortlist:
            splitter = y.split(" ")
            name = []
            for p in splitter:
                name.append(p[0].upper() + p[1:])

            upperedar.append(" ".join(name))

        sheet.insert_rows(idx=1, amount=len(messagecount)+1)
        sheet.insert_cols(idx=1, amount=2)

        sheet.cell(row=1, column=2).value = date

        for i, x in enumerate(upperedar):
            sheet.cell(row=i+2, column=1).value = x
            sheet.cell(row=i + 2, column=2).value = messagecount[x.lower()]

        workbook.save("pee.xlsx")


if __name__ == "__main__":
    main()
