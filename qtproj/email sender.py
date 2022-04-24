import mimetypes
import smtplib
import mammoth
import string
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

data_dict = {}
attachment = ""
template_file = "test.docx"
email_file = "emails.txt"
tag_files = ["bruh.txt"]

email = "emailautoformatter@gmail.com"


def remove_ws(text):
    for x in string.whitespace:
        text = text.replace(x, "")
    return text


def get_template(file_name):
    html = None
    with open(file_name, "rb") as u_file:
        if file_name[-4:] == "docx":
            text = mammoth.extract_raw_text(u_file).value
            html = mammoth.convert_to_html(u_file).value
            html = "<html><head><head><body>" + html + "<body><html>"

        else:
            text = u_file.read().decode()
        u_file.close()

    return [text, html]


def send_email(recipient, content, subject, server, attachment=None):
    message = MIMEMultipart('alternative')

    message["subject"] = subject

    if content:
        if content[1]:
            content[1] = remove_ws(content[1])
            txt = MIMEText(content[1], "html")
            message.attach(txt)
            message[""]
        elif content[0]:
            content[0] = remove_ws(content[0])
            txt = MIMEText(content[0], "plain")
            message.attach(txt)

    if attachment:
        mime_type, _ = mimetypes.guess_type(attachment)
        mime_type, mime_subtype = mime_type.split('/')
        with open(attachment, 'rb') as file:
            message.add_attachment(file.read(),
                                   maintype=mime_type,
                                   subtype=mime_subtype,
                                   filename=attachment)

    sender = email

    message["From"] = sender
    message["To"] = recipient

    server.send_message(message)


if __name__ == "__main__":
    # check format of each file to make sure that they are in the correct format
    mail_server = smtplib.SMTP_SSL("smtp.gmail.com")
    mail_server.login(email, "Bw0777$$$")
    with open(email_file, "rb") as emails:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        tags = {}
        for x in tag_files:
            file = open(x, "rb")
            tags[x.replace(".txt", "")] = [remove_ws(x.decode()) for x in file.readlines()]
        template = get_template(template_file)
        temp_let = template
        for i, email_ in enumerate(emails.readlines()):
            sheet.cell(row=i + 2, column=1).value = email_
            for p, x in enumerate(tags):
                x = remove_ws(x)
                sheet.cell(row=1, column=p + 2).value = "[" + x + "]"
                for o, thing in enumerate(temp_let):
                    if thing:
                        temp_let[o] = thing.replace("[" + x + "]", tags[x][i])
                sheet.cell(row=i + 2, column=p + 2).value = tags[x][i]
            print(temp_let)
            send_email(email_.decode(), temp_let, "temp", mail_server)
            temp_let = template
    workbook.save(filename="tester.xlsx")
    mail_server.close()
